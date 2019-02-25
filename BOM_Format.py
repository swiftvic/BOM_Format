# BOM formating 
# Original BOM 9010242321 now 123240109
import openpyxl
import re                                                                                     # Regular Expressions library
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

headers = ["Level", "Customer Part Number", "Qty", "Ref Des", "Item Rev", "Mara #", "Mara Description", "Cust MFR", "Cust MPN", "Cust Notes", "Higher Level", "Ext Qty"]

# Styling for headers
font_headers = Font(name = 'Arial', size = 10, bold = True)

# Column map from CustBOM
cust_bom_ref = {"Level": 1, "Number": 2, "BOM.Qty": 15, "BOM.Ref": 18, "Rev": 11, "Description": 4, "Manufacturers.MFR Name": 26, "Manufacturers.MPN": 27}

def main():

    filepath_old = "a_20180412_064608761.xlsx"
    filepath_new = "a_20181029_024744353.xlsx"

    wb_old = openpyxl.load_workbook(filepath_old)
    wb_new = openpyxl.load_workbook(filepath_new)

    cust_sheet = wb_old["Sheet0"]                                 # Select Sheet0 as CustBOM
    ws_new_sheet = wb_new["Sheet0"]                               # Select Sheet0 of new file as CustBOM

    mara_format = wb_old.create_sheet("PFormat")                  # Create new sheet called PFormat in old workbook
    
    # Print old wb sheet names
    print(wb_old.sheetnames)

    # Old sheet max rows and columns
    max_row = cust_sheet.max_row
    max_col = cust_sheet.max_column

    # New sheet max rows and columns
    max_new_row = ws_new_sheet.max_row
    max_new_col = ws_new_sheet.max_column

    # Prints stats of each file
    print("There are " + str(max_row) + " line items in " + str(filepath_old))
    print("There are " + str(max_new_row) + " line items in " + str(filepath_new))

    # Copies BOM lvl and Assy p/n to top
    for col in range(1,3):
        mara_format.cell(1, col).value = cust_sheet.cell(2, col).value
    
    # Copies Rev and Description to new format
    mara_format["E1"] = cust_sheet["K2"].value
    mara_format["G1"] = cust_sheet["D2"].value

    # Adding headers into new sheet
    col = 1
    for item in headers:
        mara_format.cell(2, col).value = item                                                           # Copies each header into each cell
        mara_format.cell(2, col).font = font_headers                                                    # Setting the styling of each header
        col += 1

    # Copying various customer columns (ref cust_bom_col variable at top) over to Mara new sheet
    m_col = 1                                                                                           # Start col 1 of new sheet
    for cust_col in cust_bom_ref.values():                                                              # Loop through each item in cust_bom_ref dict and use it's values
        m_row = 3                                                                                       # Start row 3 of new sheet and reset to 3 after each column is copied                                                                                     
        for cust_row in range(3, max_row):                                                              # Loop through each row in cust sheet until end         
            mara_format.cell(m_row, m_col).value = cust_sheet.cell(cust_row, cust_col).value            # Copy cells from cust sheet cell to mara format cell
            m_row += 1                                                                                  # Increase to new row of new sheet
        m_col += 1                                                                                      # Increase to next col of new sheet
        if m_col == 6 or m_col == 11:                                                                   # Skip col 6 and 11 of new sheet
            m_col += 1                                                                                  

    # Adjust columns to length of cell values
    for col in mara_format.columns:
        max_length = 0
        column = col[0].column                                                                          # Gets the Column letter/name
        for cell in col:
            try:                                                                                        # Avoid error on an empty cell
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 1)
        mara_format.column_dimensions[column].width = adjusted_width

    # Insert Higher Level
    level_ref = {}                                                                                      # Dict to hold level number ref part number
    col = 1                                                                                             # BOM level typically on column 1
    for row in range(2, max_row):                                                                       # Loop through each row of CustBOM
        try:                                                                                            # Avoid error on an empty cell
            current_level = int(cust_sheet.cell(row, col).value)                                        # Store the BOM level an int in variable current_level
            #if current_level == 0 or current_level == 1:
            level_ref[current_level] = cust_sheet.cell(row, col + 1).value                              # Store in dict "current_level : part number"
        except:                                                                                         # Just pass through if any error
            pass
    
    higher_lvl_col = 11                                                                                 # Mara formatted sheet column "Higher Level"
    col = 1                                                                                             # BOM level typically on column 1
    for row in range(3, max_new_row):
        try: 
            bom_level = int(mara_format.cell(row, col).value)
            if bom_level >= 1:                                                                          # If the Line item BOM level is >= 1
                mara_format.cell(row, higher_lvl_col).value = level_ref[bom_level - 1]                  # Ref level_ref dict for one BOM level higher
        except:
            pass 

    def check_qty(qty, ref_des):
        '''
        Input: qty, ref_des (str) \n
        Output: qty, num_ref_des, boolean qty == num_ref_des
        '''
        qty = int(qty)
        ref_des_split = ref_des.split(",")
        num_ref_des = len(ref_des_split)

        return (qty, num_ref_des, qty == num_ref_des)
 
    # Saves changes
    wb_old.save(filepath_old)

pattern_ranges = '([A-Z]+)([0-9]+)-[A-Z]+([0-9]+)'

def remove_ws(phrase):
    '''Returns all whitespaces removed in phrase'''
    
    return phrase.replace(" ", "")

def regex_ranges(string):
    pattern = '([A-Z]+)([0-9]+)-[A-Z]+([0-9]+)'                     # Regex formula to pull first Designator plus numerical range with "-"" in the middle.
    
    return re.findall(pattern, string)

def sorted_nicely(l):
    """ Sorts the given iterable in the way that is expected.
 
    Required arguments:
    l -- The iterable to be sorted.
 
    """
    convert = lambda text: int(text) if text.isdigit() else text
    alphanum_key = lambda key: [convert(c) for c in re.split('([0-9]+)', key)]
    return sorted(l, key = alphanum_key)

def unpack_des(phrase):
    
    # Variables
    breakdown_phrase = []                   # List to hold the phrase broken down into a list
    designators = ''                        # Variable to hold a string of unpacked and sorted designators
    unpacked_list = []                      # Initiate variable to hold all unpacked designator char + number

    clean_phrase = remove_ws(phrase)                                   # Removes all whitespace from phrase
    
    breakdown_phrase = clean_phrase.replace('-',',').split(',')        # Takes phrase, replace '-' w/ ',' then split using ',' Need this for union of two lists later.
  
    extracted_tuples = regex_ranges(clean_phrase)                      # Uses regex pattern to pull all ranges from phrase

     # Loops through each set of tuples of extracted designators and ranges
    for range_set in extracted_tuples:                                
        # Grabs start[1] and end[2] + 1 ranges from each tuple set, concatenates designator char with number and appends to list
        # Unpacked_list adds to itself or else each loop iteration of a tuple set will erase the variable
        unpacked_list += [range_set[0] + str(des_num) for des_num in range(int(range_set[1]), int(range_set[2]) + 1)]
    
    '''
    for range_set in extracted_tuples:                                      # Loops through each set of tuples of extracted designators and ranges
        for des_num in range(int(range_set[1]), int(range_set[2]) + 1):     # Grabs start[1] and end[2] + 1 ranges from each tuple set
            unpacked_list.append(range_set[0] + str(des_num))               # Concatenates designator char with number and appends to list
    '''
    repacked_list = list(set(unpacked_list).union(set(breakdown_phrase)))    # Takes unpacked_list and combines with single designators list

    sorted_repacked_list = sorted_nicely(repacked_list)                      # Sorts list alphanumerically using "sorted_nicely" function

    for items in sorted_repacked_list:        # Turns list back into a string with commas in between each designator
        designators += items + ','
    
    #print(designators.strip(","))    # Strips last "," at the end of string
    return designators.strip(",")     # Strips last "," at the end of string and returns string

if __name__ == '__main__':
    #main()

    new_des = unpack_des("R20, R1-R5, R9, CR1-CR4, CR8, R10-R15, CR9-CR11, MAR3-MAR5, MAR1")
    print(new_des)